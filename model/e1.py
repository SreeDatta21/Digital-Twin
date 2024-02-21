import openpyxl
import random
 
# Define variable to load the dataframe
dataframe = openpyxl.load_workbook(r'C:\Users\hp\OneDrive\Documents\Book2.xlsx')
 
# Define variable to read sheet
dataframe1 = dataframe.active
 
# Iterate the loop to read the cell valuesw32
z = {1:0.005,2:1500,3:15}
r=random.randrange(10000)
x=0
l = []
for col in dataframe1.iter_cols(1, dataframe1.max_column-1):
    l.append(col[r].value)
    x+=1
    flag=2
    if col[r].value>z[x]  :
        flag = flag - 1
        if flag == 0:
            break
if flag:
    print("y")
    print("Strain:",'{0:.2f}'.format(l[0]),"\nLoad:",'{0:.2f}'.format(l[1]),"\nDisplace:",'{0:.2f}'.format(l[2]))
    
else:
    print("n")
    print("Strain:",'{0:.2f}'.format(l[0]),"\nLoad:",'{0:.2f}'.format(l[1]),"\nDisplace:",'{0:.2f}'.format(l[2]))


