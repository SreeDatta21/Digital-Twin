import numpy as np
import joblib
import warnings
import openpyxl
import random
 
warnings.filterwarnings("ignore", category=UserWarning, module='sklearn')


model = joblib.load('trained_model.joblib')
dataframe = openpyxl.load_workbook(r'C:\Users\hp\OneDrive\Documents\Book2.xlsx')

dataframe1 = dataframe.active
l = []
r=random.randrange(10000)
for col in dataframe1.iter_cols(1, dataframe1.max_column-1):
    l.append(col[r].value)
strain = l[0]
load = l[1]
displacement = l[2]

# Make predictions using the model
user_input = np.array([[strain, load, displacement]])
result = model.predict(user_input)[0]

print(f"{result}")
print("Strain:",'{0:.2f}'.format(l[0]),"\nLoad:",'{0:.2f}'.format(l[1]),"\nDisplace:",'{0:.2f}'.format(l[2]))
